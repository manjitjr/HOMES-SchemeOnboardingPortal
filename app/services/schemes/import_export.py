import io
import json as _json
from datetime import date

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import (
    APIBatchInterfaces,
    HOMESFunctions,
    MTBands,
    SchemeMTParameters,
    SchemeMaster,
    SchemeOverview,
    SchemeSubmission,
    SubmissionStatus,
    TransactionDetails,
    User,
)


class SchemeImportExportService:
    _HDR_FONT = Font(bold=True, color="FFFFFF")
    _HDR_FILL = PatternFill("solid", fgColor="2563EB")
    _HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _REQUIRED_OVERVIEW = ["scheme_name", "agency", "scheme_code", "legislated_or_consent"]
    _REQUIRED_MT = ["same_as_applicant", "residence_status"]
    _REQUIRED_TX = ["portal_corppass", "batch_sftp", "annual_mt_apps"]
    _REQUIRED_HF = ["subscribe_notifications", "auto_mt_sub"]

    def _set_header_row(self, ws, headers: list[str]):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = self._HDR_FONT
            cell.fill = self._HDR_FILL
            cell.alignment = self._HDR_ALIGN
        ws.row_dimensions[1].height = 28

    def _sv(self, v) -> str:
        if v is None:
            return ""
        if isinstance(v, bool):
            return "Yes" if v else "No"
        if isinstance(v, (dict, list)):
            return _json.dumps(v)
        return str(v)

    def _build_scheme_excel(self, submissions: list) -> io.BytesIO:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "1. Scheme Overview"
        self._set_header_row(
            ws1,
            [
                "Scheme Name", "Agency", "Scheme Code", "Legislated/Consent", "Consent Scope",
                "Valid From", "Valid To", "Org Established", "Purpose", "Funding Source", "Governing Body",
                "Eligibility Org", "Evaluating Orgs", "Third Parties", "Group Name", "Logo Info",
            ],
        )
        for s in submissions:
            ov = s.overview
            bg = (ov.background_info or {}) if ov else {}
            ws1.append(
                [
                    self._sv(ov.scheme_name if ov else s.scheme_master_id),
                    self._sv(ov.agency if ov else ""),
                    self._sv(ov.scheme_code if ov else ""),
                    self._sv(ov.legislated_or_consent if ov else ""),
                    self._sv(ov.consent_scope if ov else ""),
                    self._sv(s.valid_from),
                    self._sv(s.valid_to),
                    self._sv(bg.get("org_established", "")),
                    self._sv(bg.get("purpose", "")),
                    self._sv(bg.get("funding_source", "")),
                    self._sv(bg.get("governing_body", "")),
                    self._sv(bg.get("eligibility_org", "")),
                    self._sv(bg.get("eval_orgs", "")),
                    self._sv(bg.get("third_parties", "")),
                    self._sv(bg.get("group_name", "")),
                    self._sv(bg.get("logo_info", "")),
                ]
            )

        ws2 = wb.create_sheet("2. Scheme MT Parameters")
        self._set_header_row(
            ws2,
            [
                "Scheme Name", "Agency", "Same As Applicant", "Relationship Desc", "Residence Status", "Foreigner Pass Types",
                "Related Used", "Related Construct", "Related Deviation", "Related Inclusion JSON",
                "Nuclear Used", "Nuclear Construct", "Nuclear Deviation", "Nuclear Inclusion JSON",
                "Parent Guardian Used", "Parent Guardian Construct", "Parent Guardian Deviation", "Parent Guardian Inclusion JSON",
                "Immediate Family Used", "Immediate Family Construct", "Immediate Family Deviation", "Immediate Family Inclusion JSON",
                "Freeform Used", "Freeform Construct", "Freeform Deviation", "Freeform Inclusion JSON",
                "Income Employment", "Income Trade", "Income Investments", "Income Rental", "Income Rollup", "AV Used", "MP Used",
            ],
        )
        for s in submissions:
            ov = s.overview
            mt = (s.mt_parameters.data or {}) if s.mt_parameters else {}
            ws2.append(
                [
                    self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else ""),
                    self._sv(mt.get("same_as_applicant", "")), self._sv(mt.get("relationship_desc", "")), self._sv(mt.get("residence_status", "")), self._sv(mt.get("foreigner_pass_types", "")),
                    self._sv(mt.get("related_used", "")), self._sv(mt.get("related_construct", "")), self._sv(mt.get("related_deviation", "")), self._sv((mt.get("related") or {}).get("inclusion", {})),
                    self._sv(mt.get("nuclear_used", "")), self._sv(mt.get("nuclear_construct", "")), self._sv(mt.get("nuclear_deviation", "")), self._sv((mt.get("nuclear") or {}).get("inclusion", {})),
                    self._sv(mt.get("parent_guardian_used", "")), self._sv(mt.get("parent_guardian_construct", "")), self._sv(mt.get("parent_guardian_deviation", "")), self._sv((mt.get("parent_guardian") or {}).get("inclusion", {})),
                    self._sv(mt.get("immediate_family_used", "")), self._sv(mt.get("immediate_family_construct", "")), self._sv(mt.get("immediate_family_deviation", "")), self._sv((mt.get("immediate_family") or {}).get("inclusion", {})),
                    self._sv(mt.get("freeform_used", "")), self._sv(mt.get("freeform_construct", "")), self._sv(mt.get("freeform_deviation", "")), self._sv((mt.get("freeform") or {}).get("inclusion", {})),
                    self._sv(mt.get("income_employment", "")), self._sv(mt.get("income_trade", "")), self._sv(mt.get("income_investments", "")), self._sv(mt.get("income_rental", "")), self._sv(mt.get("income_rollup", "")), self._sv(mt.get("av_used", "")), self._sv(mt.get("mp_used", "")),
                ]
            )

        ws3 = wb.create_sheet("3. Transaction Details")
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        self._set_header_row(
            ws3,
            [
                "Scheme Name", "Agency", "HOMES Portal (CorpPass)", "No. Orgs CorpPass", "HOMES Portal (Intranet)", "No. Orgs Intranet",
                "Batch SFTP", "Realtime APEX", "Scheduled Reports Delivery", "Required Scheduled Reports",
                "Systems via Intranet SFTP", "Systems via Intranet API", "Systems via Internet SFTP", "Systems via Internet API",
                "Interfacing System Status", "Interface Ready Date", "Interface System Names", "Annual MT Applications",
                *[f"Month {m}" for m in months],
                "Manual Recon %", "Max Concurrent Users/hr", "Reconciliation Breakdown", "SFTP Peak Vol", "SFTP Avg Vol",
                "API Peak Vol", "API Avg Vol", "Portal Peak Vol", "Portal Avg Vol", "Bulk Query Peak Vol", "Bulk Query Avg Vol",
            ],
        )
        for s in submissions:
            ov = s.overview
            tx = (s.transactions.data or {}) if s.transactions else {}
            ws3.append(
                [
                    self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else ""),
                    self._sv(tx.get("portal_corppass", "")), self._sv(tx.get("num_orgs_corppass", "")),
                    self._sv(tx.get("portal_intranet", "")), self._sv(tx.get("num_orgs_intranet", "")),
                    self._sv(tx.get("batch_sftp", "")), self._sv(tx.get("realtime_apex", "")),
                    self._sv(tx.get("sched_reports_delivery", "")), self._sv(tx.get("required_sched_reports", "")),
                    self._sv(tx.get("sys_intranet_sftp", "")), self._sv(tx.get("sys_intranet_api", "")),
                    self._sv(tx.get("sys_internet_sftp", "")), self._sv(tx.get("sys_internet_api", "")),
                    self._sv(tx.get("iface_status", "")), self._sv(tx.get("iface_ready_date", "")), self._sv(tx.get("iface_names", "")),
                    self._sv(tx.get("annual_mt_apps", "")),
                    *[self._sv(tx.get(f"month_{i + 1}", "")) for i in range(12)],
                    self._sv(tx.get("manual_recon_pct", "")), self._sv(tx.get("max_concurrent_users", "")), self._sv(tx.get("recon_breakdown", "")),
                    self._sv(tx.get("vol_sftp_peak", "")), self._sv(tx.get("vol_sftp_avg", "")),
                    self._sv(tx.get("vol_api_peak", "")), self._sv(tx.get("vol_api_avg", "")),
                    self._sv(tx.get("vol_portal_peak", "")), self._sv(tx.get("vol_portal_avg", "")),
                    self._sv(tx.get("vol_bulk_query_peak", "")), self._sv(tx.get("vol_bulk_query_avg", "")),
                ]
            )

        ws4 = wb.create_sheet("4. HOMES Functions")
        events = [
            "EV001", "EV002", "EV003", "EV004", "EV005", "EV006", "EV007",
            "EV008", "EV009", "EV010", "EV011", "EV012", "EV013", "EV014",
        ]
        self._set_header_row(
            ws4,
            [
                "Scheme Name", "Agency", "MTC View: Related", "MTC View: Nuclear", "MTC View: Parent-Guardian", "MTC View: IFM",
                "Allow Others to View?", "Can View Others?", "Result View: Related", "Result View: Nuclear", "Result View: Parent-Guardian", "Result View: IFM",
                "Affiliated Schemes", "Read MSHL?", "Subscribe Notifications?", "Beneficiary Scope",
                *[f"Event {e}" for e in events], "Auto-MT Subscription", "Cohort Basis",
            ],
        )
        mtc_list = ["related", "nuclear", "parent_guardian", "ifm"]
        for s in submissions:
            ov = s.overview
            hf = (s.homes_functions.data or {}) if s.homes_functions else {}
            ws4.append(
                [
                    self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else ""),
                    *[self._sv(hf.get(f"view_{m}", "")) for m in mtc_list],
                    self._sv(hf.get("allow_others_view", "")), self._sv(hf.get("can_view_others", "")),
                    *[self._sv(hf.get(f"result_{m}", "")) for m in mtc_list],
                    self._sv(hf.get("affiliated_schemes", "")), self._sv(hf.get("read_mshl", "")),
                    self._sv(hf.get("subscribe_notifications", "")), self._sv(hf.get("beneficiary_scope", "")),
                    *[self._sv(hf.get(f"event_{e}", "")) for e in events],
                    self._sv(hf.get("auto_mt_sub", "")), self._sv(hf.get("cohort_basis", "")),
                ]
            )

        ws5 = wb.create_sheet("5. MT Bands")
        self._set_header_row(
            ws5,
            [
                "Scheme Name", "Agency", "Row Type", "Band Name", "Band Version", "GI Subsidy Formula", "PCI Subsidy Formula",
                "Min Income", "Max Income", "Min AV", "Max AV", "Ranking Label", "Ranking Order",
            ],
        )
        for s in submissions:
            ov = s.overview
            mb = (s.mt_bands.data or {}) if s.mt_bands else {}
            bands = mb.get("bands", []) if isinstance(mb, dict) else []
            rankings = mb.get("rankings", []) if isinstance(mb, dict) else []
            if not bands and not rankings:
                ws5.append([self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else ""), "", "", "", "", "", "", "", "", "", "", ""])
            for b in bands:
                ws5.append(
                    [
                        self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else ""), "Band",
                        self._sv(b.get("band_name", "")), self._sv(b.get("band_version", "")), self._sv(b.get("gi_formula", "")), self._sv(b.get("pci_formula", "")),
                        self._sv(b.get("min_income", "")), self._sv(b.get("max_income", "")), self._sv(b.get("min_av", "")), self._sv(b.get("max_av", "")), "", "",
                    ]
                )
            for r in rankings:
                ws5.append(
                    [
                        self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else ""), "Ranking",
                        "", "", "", "", "", "", "", "", self._sv(r.get("label", "")), self._sv(r.get("order", "")),
                    ]
                )

        ws6 = wb.create_sheet("6. API & Batch Interfaces")
        api_keys = [*[f"P12 API{i}" for i in range(1, 20)], "P20 API"]
        hdr6 = ["Scheme Name", "Agency"]
        for k in api_keys:
            hdr6 += [f"{k} Used", f"{k} Avg TPS", f"{k} Peak TPS"]
        self._set_header_row(ws6, hdr6)
        for s in submissions:
            ov = s.overview
            ai = (s.api_interfaces.data or {}) if s.api_interfaces else {}
            row6 = [self._sv(ov.scheme_name if ov else ""), self._sv(ov.agency if ov else "")]
            for i in range(1, 20):
                row6 += [self._sv(ai.get(f"api_P12_API{i}_used", "")), self._sv(ai.get(f"api_P12_API{i}_avg", "")), self._sv(ai.get(f"api_P12_API{i}_peak", ""))]
            row6 += [self._sv(ai.get("api_P20_API_used", "")), self._sv(ai.get("api_P20_API_avg", "")), self._sv(ai.get("api_P20_API_peak", ""))]
            ws6.append(row6)

        for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _validate_imported_scheme(self, tab_data: dict) -> list[str]:
        warnings = []
        ov = tab_data.get("overview", {})
        for field_name in self._REQUIRED_OVERVIEW:
            if not ov.get(field_name):
                warnings.append(f"Overview: '{field_name}' is required")

        mt = tab_data.get("mt_parameters", {})
        for field_name in self._REQUIRED_MT:
            if not mt.get(field_name):
                warnings.append(f"MT Parameters: '{field_name}' is required")

        tx = tab_data.get("transactions", {})
        for field_name in self._REQUIRED_TX:
            if not tx.get(field_name):
                warnings.append(f"Transaction Details: '{field_name}' is required")

        hf = tab_data.get("homes_functions", {})
        for field_name in self._REQUIRED_HF:
            if not hf.get(field_name):
                warnings.append(f"HOMES Functions: '{field_name}' is required")

        mb = tab_data.get("mt_bands", {})
        if not mb.get("bands"):
            warnings.append("MT Bands: at least one band row is required")

        return warnings

    async def export_bulk(self, db: AsyncSession, user: User):
        query = (
            select(SchemeSubmission)
            .options(
                selectinload(SchemeSubmission.overview),
                selectinload(SchemeSubmission.mt_parameters),
                selectinload(SchemeSubmission.transactions),
                selectinload(SchemeSubmission.homes_functions),
                selectinload(SchemeSubmission.mt_bands),
                selectinload(SchemeSubmission.api_interfaces),
            )
            .order_by(SchemeSubmission.created_at.desc())
        )
        if not user.is_admin() and user.agency:
            query = query.join(SchemeOverview, SchemeSubmission.scheme_overview_id == SchemeOverview.id).where(
                SchemeOverview.agency == user.agency
            )

        result = await db.execute(query)
        submissions = result.scalars().all()
        buf = self._build_scheme_excel(submissions)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="schemes_export.xlsx"'},
        )

    async def import_schemes(self, db: AsyncSession, user: User, file: UploadFile):
        if not file.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are supported")

        content = await file.read()
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

        def _ws_rows(*sheet_names: str):
            ws = None
            for n in sheet_names:
                if n in wb.sheetnames:
                    ws = wb[n]
                    break
            if ws is None:
                return []
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).strip() if h else "" for h in rows[0]]
            return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]

        def _v(row, *keys):
            for k in keys:
                if k in row and row[k] not in (None, ""):
                    return str(row[k]).strip()
            return ""

        overview_rows = _ws_rows("1. Scheme Overview", "Scheme Overview")
        schemes_map: dict[tuple, dict] = {}
        for row in overview_rows:
            name = _v(row, "Scheme Name")
            agency = _v(row, "Agency")
            if not name:
                continue
            key = (name, agency)
            schemes_map[key] = {
                "overview": {
                    "scheme_name": name,
                    "agency": agency,
                    "scheme_code": _v(row, "Scheme Code"),
                    "legislated_or_consent": _v(row, "Legislated/Consent"),
                    "consent_scope": _v(row, "Consent Scope"),
                    "valid_from": _v(row, "Valid From"),
                    "valid_to": _v(row, "Valid To"),
                    "background_info": {
                        "org_established": _v(row, "Org Established"),
                        "purpose": _v(row, "Purpose"),
                        "funding_source": _v(row, "Funding Source"),
                        "governing_body": _v(row, "Governing Body"),
                        "eligibility_org": _v(row, "Eligibility Org"),
                        "eval_orgs": _v(row, "Evaluating Orgs"),
                        "third_parties": _v(row, "Third Parties"),
                        "group_name": _v(row, "Group Name"),
                        "logo_info": _v(row, "Logo Info"),
                    },
                },
                "mt_parameters": {},
                "transactions": {},
                "homes_functions": {},
                "mt_bands": {"bands": [], "rankings": []},
                "api_interfaces": {},
            }

        for row in _ws_rows("2. Scheme MT Parameters", "Scheme MT Parameters"):
            key = (_v(row, "Scheme Name"), _v(row, "Agency"))
            if key not in schemes_map:
                continue

            def _inclusion(col):
                raw = _v(row, col)
                if not raw:
                    return {}
                try:
                    parsed = _json.loads(raw)
                    return parsed if isinstance(parsed, dict) else {}
                except Exception:
                    return {}

            schemes_map[key]["mt_parameters"] = {
                "same_as_applicant": _v(row, "Same As Applicant"),
                "relationship_desc": _v(row, "Relationship Desc"),
                "residence_status": _v(row, "Residence Status"),
                "foreigner_pass_types": _v(row, "Foreigner Pass Types"),
                "related_used": _v(row, "Related Used"),
                "related_construct": _v(row, "Related Construct"),
                "related_deviation": _v(row, "Related Deviation"),
                "related": {"inclusion": _inclusion("Related Inclusion JSON")},
                "nuclear_used": _v(row, "Nuclear Used"),
                "nuclear_construct": _v(row, "Nuclear Construct"),
                "nuclear_deviation": _v(row, "Nuclear Deviation"),
                "nuclear": {"inclusion": _inclusion("Nuclear Inclusion JSON")},
                "parent_guardian_used": _v(row, "Parent Guardian Used"),
                "parent_guardian_construct": _v(row, "Parent Guardian Construct"),
                "parent_guardian_deviation": _v(row, "Parent Guardian Deviation"),
                "parent_guardian": {"inclusion": _inclusion("Parent Guardian Inclusion JSON")},
                "immediate_family_used": _v(row, "Immediate Family Used"),
                "immediate_family_construct": _v(row, "Immediate Family Construct"),
                "immediate_family_deviation": _v(row, "Immediate Family Deviation"),
                "immediate_family": {"inclusion": _inclusion("Immediate Family Inclusion JSON")},
                "freeform_used": _v(row, "Freeform Used"),
                "freeform_construct": _v(row, "Freeform Construct"),
                "freeform_deviation": _v(row, "Freeform Deviation"),
                "freeform": {"inclusion": _inclusion("Freeform Inclusion JSON")},
                "income_employment": _v(row, "Income Employment"),
                "income_trade": _v(row, "Income Trade"),
                "income_investments": _v(row, "Income Investments"),
                "income_rental": _v(row, "Income Rental"),
                "income_rollup": _v(row, "Income Rollup"),
                "av_used": _v(row, "AV Used"),
                "mp_used": _v(row, "MP Used"),
            }

        months_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        for row in _ws_rows("3. Transaction Details", "Transaction Details"):
            key = (_v(row, "Scheme Name"), _v(row, "Agency"))
            if key not in schemes_map:
                continue
            tx: dict = {
                "portal_corppass": _v(row, "HOMES Portal (CorpPass)"),
                "num_orgs_corppass": _v(row, "No. Orgs CorpPass"),
                "portal_intranet": _v(row, "HOMES Portal (Intranet)"),
                "num_orgs_intranet": _v(row, "No. Orgs Intranet"),
                "batch_sftp": _v(row, "Batch SFTP"),
                "realtime_apex": _v(row, "Realtime APEX"),
                "sched_reports_delivery": _v(row, "Scheduled Reports Delivery"),
                "required_sched_reports": _v(row, "Required Scheduled Reports"),
                "sys_intranet_sftp": _v(row, "Systems via Intranet SFTP"),
                "sys_intranet_api": _v(row, "Systems via Intranet API"),
                "sys_internet_sftp": _v(row, "Systems via Internet SFTP"),
                "sys_internet_api": _v(row, "Systems via Internet API"),
                "iface_status": _v(row, "Interfacing System Status"),
                "iface_ready_date": _v(row, "Interface Ready Date"),
                "iface_names": _v(row, "Interface System Names"),
                "annual_mt_apps": _v(row, "Annual MT Applications"),
                "manual_recon_pct": _v(row, "Manual Recon %"),
                "max_concurrent_users": _v(row, "Max Concurrent Users/hr"),
                "recon_breakdown": _v(row, "Reconciliation Breakdown"),
                "vol_sftp_peak": _v(row, "SFTP Peak Vol"),
                "vol_sftp_avg": _v(row, "SFTP Avg Vol"),
                "vol_api_peak": _v(row, "API Peak Vol"),
                "vol_api_avg": _v(row, "API Avg Vol"),
                "vol_portal_peak": _v(row, "Portal Peak Vol"),
                "vol_portal_avg": _v(row, "Portal Avg Vol"),
                "vol_bulk_query_peak": _v(row, "Bulk Query Peak Vol"),
                "vol_bulk_query_avg": _v(row, "Bulk Query Avg Vol"),
            }
            for i, m in enumerate(months_abbr):
                tx[f"month_{i + 1}"] = _v(row, f"Month {m}")
            schemes_map[key]["transactions"] = tx

        hf_events = [f"EV{str(i).zfill(3)}" for i in range(1, 15)]
        hf_mtcs = ["related", "nuclear", "parent_guardian", "ifm"]
        for row in _ws_rows("4. HOMES Functions", "HOMES Functions"):
            key = (_v(row, "Scheme Name"), _v(row, "Agency"))
            if key not in schemes_map:
                continue
            hf: dict = {
                "allow_others_view": _v(row, "Allow Others to View?"),
                "can_view_others": _v(row, "Can View Others?"),
                "affiliated_schemes": _v(row, "Affiliated Schemes"),
                "read_mshl": _v(row, "Read MSHL?"),
                "subscribe_notifications": _v(row, "Subscribe Notifications?"),
                "beneficiary_scope": _v(row, "Beneficiary Scope"),
                "auto_mt_sub": _v(row, "Auto-MT Subscription"),
                "cohort_basis": _v(row, "Cohort Basis"),
            }
            for m in hf_mtcs:
                label = "IFM" if m == "ifm" else m.replace("_", "-").title()
                hf[f"view_{m}"] = _v(row, f"MTC View: {label}")
                hf[f"result_{m}"] = _v(row, f"Result View: {label}")
            for e in hf_events:
                hf[f"event_{e}"] = _v(row, f"Event {e}")
            schemes_map[key]["homes_functions"] = hf

        for row in _ws_rows("5. MT Bands", "MT Bands"):
            key = (_v(row, "Scheme Name"), _v(row, "Agency"))
            if key not in schemes_map:
                continue
            row_type = _v(row, "Row Type")
            if row_type == "Band":
                schemes_map[key]["mt_bands"]["bands"].append(
                    {
                        "band_name": _v(row, "Band Name"),
                        "band_version": _v(row, "Band Version"),
                        "gi_formula": _v(row, "GI Subsidy Formula"),
                        "pci_formula": _v(row, "PCI Subsidy Formula"),
                        "min_income": _v(row, "Min Income"),
                        "max_income": _v(row, "Max Income"),
                        "min_av": _v(row, "Min AV"),
                        "max_av": _v(row, "Max AV"),
                    }
                )
            elif row_type == "Ranking":
                schemes_map[key]["mt_bands"]["rankings"].append(
                    {"label": _v(row, "Ranking Label"), "order": _v(row, "Ranking Order")}
                )

        for row in _ws_rows("6. API & Batch Interfaces", "API & Batch Interfaces"):
            key = (_v(row, "Scheme Name"), _v(row, "Agency"))
            if key not in schemes_map:
                continue
            ai: dict = {}
            for i in range(1, 20):
                api_key = f"P12 API{i}"
                ai[f"api_P12_API{i}_used"] = _v(row, f"{api_key} Used")
                ai[f"api_P12_API{i}_avg"] = _v(row, f"{api_key} Avg TPS")
                ai[f"api_P12_API{i}_peak"] = _v(row, f"{api_key} Peak TPS")
            ai["api_P20_API_used"] = _v(row, "P20 API Used")
            ai["api_P20_API_avg"] = _v(row, "P20 API Avg TPS")
            ai["api_P20_API_peak"] = _v(row, "P20 API Peak TPS")
            schemes_map[key]["api_interfaces"] = ai

        agency_filter = None if user.is_admin() else user.agency
        results = []
        for (name, agency), tab_data in schemes_map.items():
            if agency_filter and agency != agency_filter:
                results.append(
                    {
                        "scheme_name": name,
                        "agency": agency,
                        "status": "skipped",
                        "reason": "Agency mismatch — you can only import into your own agency",
                        "warnings": [],
                    }
                )
                continue

            ov_data = tab_data["overview"]
            existing_q = await db.execute(
                select(SchemeSubmission)
                .join(SchemeOverview, SchemeSubmission.scheme_overview_id == SchemeOverview.id)
                .where(
                    SchemeOverview.scheme_name == name,
                    SchemeOverview.agency == agency,
                    SchemeSubmission.status == SubmissionStatus.draft.value,
                )
                .options(
                    selectinload(SchemeSubmission.overview),
                    selectinload(SchemeSubmission.mt_parameters),
                    selectinload(SchemeSubmission.transactions),
                    selectinload(SchemeSubmission.homes_functions),
                    selectinload(SchemeSubmission.mt_bands),
                    selectinload(SchemeSubmission.api_interfaces),
                )
                .limit(1)
            )
            sub = existing_q.scalars().first()
            imported_new = sub is None

            if sub is None:
                master = SchemeMaster(
                    agency=agency or "",
                    scheme_name=name,
                    scheme_code=ov_data.get("scheme_code"),
                    created_by=user.id,
                )
                db.add(master)
                await db.flush()

                ov = SchemeOverview(
                    agency=agency,
                    scheme_name=name,
                    scheme_code=ov_data.get("scheme_code"),
                    legislated_or_consent=ov_data.get("legislated_or_consent"),
                    consent_scope=ov_data.get("consent_scope"),
                    background_info=ov_data.get("background_info", {}),
                )
                db.add(ov)
                await db.flush()

                valid_from = None
                valid_to = None
                try:
                    if ov_data.get("valid_from"):
                        valid_from = date.fromisoformat(str(ov_data["valid_from"])[:10])
                    if ov_data.get("valid_to"):
                        valid_to = date.fromisoformat(str(ov_data["valid_to"])[:10])
                except Exception:
                    pass

                sub = SchemeSubmission(
                    scheme_master_id=master.id,
                    scheme_overview_id=ov.id,
                    status=SubmissionStatus.draft.value,
                    version=1,
                    valid_from=valid_from,
                    valid_to=valid_to,
                    created_by=user.id,
                )
                db.add(sub)
                await db.flush()
            else:
                ov = sub.overview
                if ov:
                    ov.scheme_code = ov_data.get("scheme_code") or ov.scheme_code
                    ov.legislated_or_consent = ov_data.get("legislated_or_consent") or ov.legislated_or_consent
                    ov.consent_scope = ov_data.get("consent_scope") or ov.consent_scope
                    ov.background_info = ov_data.get("background_info") or ov.background_info

            async def _upsert_tab(model_cls, data_key: str):
                obj_res = await db.execute(select(model_cls).where(model_cls.submission_id == sub.id).limit(1))
                obj = obj_res.scalars().first()
                data = tab_data.get(data_key, {})
                if obj is None:
                    obj = model_cls(submission_id=sub.id, data=data)
                    db.add(obj)
                else:
                    obj.data = data

            await _upsert_tab(SchemeMTParameters, "mt_parameters")
            await _upsert_tab(TransactionDetails, "transactions")
            await _upsert_tab(HOMESFunctions, "homes_functions")
            await _upsert_tab(MTBands, "mt_bands")
            await _upsert_tab(APIBatchInterfaces, "api_interfaces")

            results.append(
                {
                    "scheme_name": name,
                    "agency": agency,
                    "status": "created" if imported_new else "updated",
                    "warnings": self._validate_imported_scheme(tab_data),
                }
            )

        await db.commit()
        return {
            "imported": len([r for r in results if r["status"] in ("created", "updated")]),
            "results": results,
        }
