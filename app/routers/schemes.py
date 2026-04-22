import uuid
import io, json as _json
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.config import get_db
from app.auth import get_current_user
from app.config import settings
from app.models import (
    User, SchemeOverview, SchemeSubmission, SchemeMTParameters,
    TransactionDetails, HOMESFunctions, MTBands, APIBatchInterfaces,
    ChangeLog, Comment, SubmissionStatus, OnboardingSlot, SchemeMaster, NotificationLog, sg_now,
)
from app.services.notifications import send_email_notification

router = APIRouter(prefix="/api/schemes", tags=["schemes"])

# ── Pydantic schemas ────────────────────────────────────────────

class SchemeCreate(BaseModel):
    agency: str | None = None
    scheme_name: str
    scheme_code: str | None = None
    valid_from: str | None = None
    valid_to: str | None = None
    legislated_or_consent: str | None = None
    consent_scope: str | None = None
    background_info: dict | None = None

class SchemeUpdate(BaseModel):
    agency: str | None = None
    scheme_name: str | None = None
    scheme_code: str | None = None
    valid_from: str | None = None
    valid_to: str | None = None
    legislated_or_consent: str | None = None
    consent_scope: str | None = None
    background_info: dict | None = None


class CloneVersionBody(BaseModel):
    valid_from: str
    valid_to: str


class RetireBody(BaseModel):
    retire_date: str | None = None
    comment: str | None = None

class TabUpdate(BaseModel):
    data: dict

class CommentCreate(BaseModel):
    text: str
    stage: str | None = None

class RejectBody(BaseModel):
    comment: str | None = None

class SlotSelection(BaseModel):
    year: int
    slot_month: int  # 1=Jan, 4=Apr, 7=Jul, 11=Nov
    technical_go_live: str  # ISO date string YYYY-MM-DD
    business_go_live: str  # ISO date string YYYY-MM-DD
    justification: str | None = None  # Required for is_additional=True

# ── Helpers ──────────────────────────────────────────────────────

TAB_MODEL_MAP = {
    "mt_parameters": SchemeMTParameters,
    "transactions": TransactionDetails,
    "homes_functions": HOMESFunctions,
    "mt_bands": MTBands,
    "api_interfaces": APIBatchInterfaces,
}

TAB_REL_MAP = {
    "mt_parameters": "mt_parameters",
    "transactions": "transactions",
    "homes_functions": "homes_functions",
    "mt_bands": "mt_bands",
    "api_interfaces": "api_interfaces",
}


def _require_role(user: User, *roles: str):
    if not user.has_role(*roles):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=f"Roles {user.roles} not allowed")


def _assert_agency_access(user: User, sub: SchemeSubmission):
    if user.is_admin():
        return
    if not sub.overview or not user.agency or sub.overview.agency != user.agency:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied for this agency")


def _can_edit_submission(user: User, sub: SchemeSubmission) -> bool:
    if user.has_role("mto_admin"):
        return True
    status_val = sub.status
    if user.has_role("agency_creator") and status_val in (SubmissionStatus.draft.value, SubmissionStatus.rejected.value):
        return True
    if user.has_role("agency_approver") and status_val in (SubmissionStatus.pending_review.value, SubmissionStatus.rejected.value):
        return True
    return False


def _can_edit_primary_slot(user: User, sub: SchemeSubmission, primary_slot: OnboardingSlot | None) -> bool:
    if _can_edit_submission(user, sub):
        return True
    if user.has_role("agency_creator") and primary_slot and primary_slot.approval_status == "rejected":
        return True
    return False


def _parse_iso_date(value: str | None, field_name: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{field_name} must be in YYYY-MM-DD format")


def _effective_status(sub: SchemeSubmission) -> str:
    if sub.status == SubmissionStatus.active.value and sub.valid_to and sub.valid_to < date.today():
        return SubmissionStatus.expired.value
    return sub.status


async def _approval_recipients(db: AsyncSession, role: str, agency: str | None = None) -> list[str]:
    result = await db.execute(select(User).where(User.is_active == True))
    users = result.scalars().all()
    recipients: list[str] = []
    for u in users:
        if not u.has_role(role):
            continue
        if role == "agency_approver" and agency and u.agency != agency:
            continue
        if u.email:
            recipients.append(u.email)
    return recipients


def _build_review_email(sub: SchemeSubmission, stage: str) -> tuple[str, str, str]:
    scheme_name = (sub.overview.scheme_name if sub.overview else "Unnamed Scheme")
    agency = (sub.overview.agency if sub.overview else "-")
    scheme_code = (sub.overview.scheme_code if sub.overview else "-")
    detail_url = f"{settings.app_base_url}/"

    if stage == SubmissionStatus.pending_review.value:
        subject = f"[HOMES] Scheme Pending Review: {scheme_name}"
        action = "A scheme has been submitted and is now pending agency review."
    else:
        subject = f"[HOMES] Scheme Pending Final Approval: {scheme_name}"
        action = "A scheme has been approved by agency approver and is now pending final MTO approval."

    text = (
        f"{action}\n\n"
        f"Scheme: {scheme_name}\n"
        f"Scheme Code: {scheme_code}\n"
        f"Agency: {agency}\n"
        f"Submission ID: {sub.id}\n\n"
        f"Open in portal: {detail_url}\n"
    )

    html = (
        f"<p>{action}</p>"
        f"<ul>"
        f"<li><strong>Scheme:</strong> {scheme_name}</li>"
        f"<li><strong>Scheme Code:</strong> {scheme_code}</li>"
        f"<li><strong>Agency:</strong> {agency}</li>"
        f"<li><strong>Submission ID:</strong> {sub.id}</li>"
        f"</ul>"
        f"<p><a href=\"{detail_url}\">Open in HOMES Portal</a></p>"
    )
    return subject, text, html


async def _notify_for_workflow_stage(db: AsyncSession, sub: SchemeSubmission, stage: str, triggered_by: str | None = None):
    if stage == SubmissionStatus.pending_review.value:
        recipients = await _approval_recipients(db, "agency_approver", agency=(sub.overview.agency if sub.overview else None))
    elif stage == SubmissionStatus.pending_final.value:
        recipients = await _approval_recipients(db, "mto_admin")
    else:
        return

    subject, text, html = _build_review_email(sub, stage)
    result = send_email_notification(recipients, subject, text, html)
    db.add(
        NotificationLog(
            submission_id=sub.id,
            stage=stage,
            subject=subject,
            recipients=result.get("recipients") or recipients,
            delivery_status=result.get("status") or "skipped",
            detail=result.get("detail"),
            triggered_by=triggered_by,
        )
    )
    await db.commit()


@router.get("/notification-logs")
async def list_notification_logs(
    limit: int = 200,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _require_role(user, "mto_admin")
    safe_limit = max(1, min(limit, 1000))

    result = await db.execute(
        select(NotificationLog).order_by(NotificationLog.created_at.desc()).limit(safe_limit)
    )
    logs = result.scalars().all()

    out = []
    for log in logs:
        sub = await db.get(SchemeSubmission, log.submission_id) if log.submission_id else None
        ov = await db.get(SchemeOverview, sub.scheme_overview_id) if sub and sub.scheme_overview_id else None
        actor = await db.get(User, log.triggered_by) if log.triggered_by else None
        out.append(
            {
                "id": log.id,
                "submission_id": log.submission_id,
                "scheme_name": ov.scheme_name if ov else None,
                "agency": ov.agency if ov else None,
                "stage": log.stage,
                "subject": log.subject,
                "recipients": log.recipients or [],
                "delivery_status": log.delivery_status,
                "detail": log.detail,
                "triggered_by": actor.display_name if actor else None,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
        )
    return out


async def _ensure_master_for_submission(db: AsyncSession, sub: SchemeSubmission, user: User | None = None):
    if sub.scheme_master_id:
        return
    ov = sub.overview
    if not ov:
        return
    m = SchemeMaster(
        agency=ov.agency or "",
        scheme_name=ov.scheme_name,
        scheme_code=ov.scheme_code,
        created_by=(user.id if user else sub.created_by),
    )
    db.add(m)
    await db.flush()
    sub.scheme_master_id = m.id


async def _validate_version_window(
    db: AsyncSession,
    master_id: str,
    valid_from: date | None,
    valid_to: date | None,
    *,
    exclude_submission_id: str | None = None,
):
    if not valid_from or not valid_to:
        return
    if valid_from > valid_to:
        raise HTTPException(status_code=400, detail="valid_from cannot be after valid_to")

    res = await db.execute(
        select(SchemeSubmission).where(SchemeSubmission.scheme_master_id == master_id)
    )
    versions = res.scalars().all()

    for v in versions:
        if exclude_submission_id and v.id == exclude_submission_id:
            continue
        if not v.valid_from or not v.valid_to:
            continue
        if not (valid_to < v.valid_from or valid_from > v.valid_to):
            raise HTTPException(
                status_code=400,
                detail=f"Version date range overlaps with version v{v.version} ({v.valid_from} to {v.valid_to})",
            )

    active_versions = [v for v in versions if _effective_status(v) == SubmissionStatus.active.value and (not exclude_submission_id or v.id != exclude_submission_id)]
    if active_versions:
        active = active_versions[0]
        if not active.valid_to:
            raise HTTPException(
                status_code=400,
                detail=f"Current active version v{active.version} has no end date. Set valid_to before creating a new version.",
            )
        if active.valid_to and valid_from <= active.valid_to:
            raise HTTPException(
                status_code=400,
                detail=f"New version cannot start before active version v{active.version} ends on {active.valid_to}",
            )


async def _get_submission(db: AsyncSession, scheme_id: str) -> SchemeSubmission:
    result = await db.execute(
        select(SchemeSubmission)
        .options(
            selectinload(SchemeSubmission.master),
            selectinload(SchemeSubmission.overview),
            selectinload(SchemeSubmission.mt_parameters),
            selectinload(SchemeSubmission.transactions),
            selectinload(SchemeSubmission.homes_functions),
            selectinload(SchemeSubmission.mt_bands),
            selectinload(SchemeSubmission.api_interfaces),
            selectinload(SchemeSubmission.creator),
            selectinload(SchemeSubmission.onboarding_slots),
        )
        .where(SchemeSubmission.id == scheme_id)
    )
    sub = result.scalar_one_or_none()
    if not sub:
        raise HTTPException(status_code=404, detail="Scheme not found")
    return sub


def _sub_to_dict(sub: SchemeSubmission) -> dict:
    ov = sub.overview
    master = sub.master
    eff_status = _effective_status(sub)
    slots = []
    if sub.onboarding_slots:
        for slot in sorted(sub.onboarding_slots, key=lambda x: (x.is_additional, x.slot_month)):
            slots.append({
                "id": slot.id,
                "year": slot.year,
                "slot_month": slot.slot_month,
                "slot_month_name": slot.slot_month_name,
                "is_additional": slot.is_additional,
                "technical_go_live": slot.technical_go_live.isoformat() if slot.technical_go_live else None,
                "business_go_live": slot.business_go_live.isoformat() if slot.business_go_live else None,
                "justification": slot.justification,
                "approval_status": slot.approval_status,
                "approver_comment": slot.approver_comment,
                "booked_at": slot.booked_at.isoformat() if slot.booked_at else None,
            })
    
    return {
        "id": sub.id,
        "scheme_master_id": sub.scheme_master_id,
        "agency": (master.agency if master else (ov.agency if ov else None)),
        "scheme_name": (master.scheme_name if master else (ov.scheme_name if ov else None)),
        "scheme_code": (master.scheme_code if master else (ov.scheme_code if ov else None)),
        "legislated_or_consent": ov.legislated_or_consent if ov else None,
        "consent_scope": ov.consent_scope if ov else None,
        "background_info": ov.background_info if ov else None,
        "status": eff_status,
        "version": sub.version,
        "version_label": f"v{sub.version}",
        "valid_from": sub.valid_from.isoformat() if sub.valid_from else None,
        "valid_to": sub.valid_to.isoformat() if sub.valid_to else None,
        "cloned_from_submission_id": sub.cloned_from_submission_id,
        "created_by": sub.creator.display_name if sub.creator else None,
        "created_at": sub.created_at.isoformat() if sub.created_at else None,
        "updated_at": sub.updated_at.isoformat() if sub.updated_at else None,
        "overview": {
            "id": ov.id,
            "agency": ov.agency,
            "scheme_name": ov.scheme_name,
            "scheme_code": ov.scheme_code,
            "legislated_or_consent": ov.legislated_or_consent,
            "consent_scope": ov.consent_scope,
            "background_info": ov.background_info,
        } if ov else None,
        "onboarding_slots": slots,
        "mt_parameters": sub.mt_parameters.data if sub.mt_parameters else None,
        "transactions": sub.transactions.data if sub.transactions else None,
        "homes_functions": sub.homes_functions.data if sub.homes_functions else None,
        "mt_bands": sub.mt_bands.data if sub.mt_bands else None,
        "api_interfaces": sub.api_interfaces.data if sub.api_interfaces else None,
    }


_MISSING = object()


def _json_safe(value):
    if value is _MISSING:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    return value


def _deep_field_diff(old_value, new_value, path: str = "") -> list[dict]:
    changes: list[dict] = []

    if isinstance(old_value, dict) and isinstance(new_value, dict):
        for key in sorted(set(old_value.keys()) | set(new_value.keys())):
            o = old_value.get(key, _MISSING)
            n = new_value.get(key, _MISSING)
            field_path = f"{path}.{key}" if path else str(key)
            changes.extend(_deep_field_diff(o, n, field_path))
        return changes

    if isinstance(old_value, list) and isinstance(new_value, list):
        max_len = max(len(old_value), len(new_value))
        for idx in range(max_len):
            o = old_value[idx] if idx < len(old_value) else _MISSING
            n = new_value[idx] if idx < len(new_value) else _MISSING
            field_path = f"{path}[{idx}]" if path else f"[{idx}]"
            changes.extend(_deep_field_diff(o, n, field_path))
        return changes

    if old_value is _MISSING and new_value is _MISSING:
        return changes
    if old_value == new_value:
        return changes

    if old_value is _MISSING:
        change_type = "added"
    elif new_value is _MISSING:
        change_type = "removed"
    else:
        change_type = "updated"

    changes.append({
        "field": path or "value",
        "old": _json_safe(old_value),
        "new": _json_safe(new_value),
        "change_type": change_type,
    })
    return changes

# ── CRUD ─────────────────────────────────────────────────────────

# ── Bulk Export / Import ─────────────────────────────────────────

# ─── helpers for structured Excel export ────────────────────────

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="2563EB")  # blue-600
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_header_row(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _HDR_ALIGN
    ws.row_dimensions[1].height = 28


def _sv(v) -> str:
    """Safe string value for Excel cells."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (dict, list)):
        return _json.dumps(v)
    return str(v)


def _build_scheme_excel(submissions: list) -> io.BytesIO:
    """Build structured multi-sheet Excel matching the 6-tab format."""
    wb = Workbook()

    # ── Sheet 1: Scheme Overview ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "1. Scheme Overview"
    _set_header_row(ws1, [
        "Scheme Name", "Agency", "Scheme Code", "Legislated/Consent", "Consent Scope",
        "Valid From", "Valid To",
        "Org Established", "Purpose", "Funding Source", "Governing Body",
        "Eligibility Org", "Evaluating Orgs", "Third Parties", "Group Name", "Logo Info",
    ])
    for s in submissions:
        ov = s.overview
        bg = (ov.background_info or {}) if ov else {}
        ws1.append([
            _sv(ov.scheme_name if ov else s.scheme_master_id),
            _sv(ov.agency if ov else ""),
            _sv(ov.scheme_code if ov else ""),
            _sv(ov.legislated_or_consent if ov else ""),
            _sv(ov.consent_scope if ov else ""),
            _sv(s.valid_from),
            _sv(s.valid_to),
            _sv(bg.get("org_established", "")),
            _sv(bg.get("purpose", "")),
            _sv(bg.get("funding_source", "")),
            _sv(bg.get("governing_body", "")),
            _sv(bg.get("eligibility_org", "")),
            _sv(bg.get("eval_orgs", "")),
            _sv(bg.get("third_parties", "")),
            _sv(bg.get("group_name", "")),
            _sv(bg.get("logo_info", "")),
        ])

    # ── Sheet 2: Scheme MT Parameters ────────────────────────────-
    ws2 = wb.create_sheet("2. Scheme MT Parameters")
    _set_header_row(ws2, [
        "Scheme Name", "Agency",
        "Same As Applicant", "Relationship Desc", "Residence Status", "Foreigner Pass Types",
        "Related Used", "Related Construct", "Related Deviation",
        "Related Inclusion JSON",
        "Nuclear Used", "Nuclear Construct", "Nuclear Deviation",
        "Nuclear Inclusion JSON",
        "Parent Guardian Used", "Parent Guardian Construct", "Parent Guardian Deviation",
        "Parent Guardian Inclusion JSON",
        "Immediate Family Used", "Immediate Family Construct", "Immediate Family Deviation",
        "Immediate Family Inclusion JSON",
        "Freeform Used", "Freeform Construct", "Freeform Deviation",
        "Freeform Inclusion JSON",
        "Income Employment", "Income Trade", "Income Investments", "Income Rental", "Income Rollup",
        "AV Used", "MP Used",
    ])
    for s in submissions:
        ov = s.overview
        mt = (s.mt_parameters.data or {}) if s.mt_parameters else {}
        ws2.append([
            _sv(ov.scheme_name if ov else ""),
            _sv(ov.agency if ov else ""),
            _sv(mt.get("same_as_applicant", "")),
            _sv(mt.get("relationship_desc", "")),
            _sv(mt.get("residence_status", "")),
            _sv(mt.get("foreigner_pass_types", "")),
            _sv(mt.get("related_used", "")), _sv(mt.get("related_construct", "")), _sv(mt.get("related_deviation", "")),
            _sv((mt.get("related") or {}).get("inclusion", {})),
            _sv(mt.get("nuclear_used", "")), _sv(mt.get("nuclear_construct", "")), _sv(mt.get("nuclear_deviation", "")),
            _sv((mt.get("nuclear") or {}).get("inclusion", {})),
            _sv(mt.get("parent_guardian_used", "")), _sv(mt.get("parent_guardian_construct", "")), _sv(mt.get("parent_guardian_deviation", "")),
            _sv((mt.get("parent_guardian") or {}).get("inclusion", {})),
            _sv(mt.get("immediate_family_used", "")), _sv(mt.get("immediate_family_construct", "")), _sv(mt.get("immediate_family_deviation", "")),
            _sv((mt.get("immediate_family") or {}).get("inclusion", {})),
            _sv(mt.get("freeform_used", "")), _sv(mt.get("freeform_construct", "")), _sv(mt.get("freeform_deviation", "")),
            _sv((mt.get("freeform") or {}).get("inclusion", {})),
            _sv(mt.get("income_employment", "")), _sv(mt.get("income_trade", "")),
            _sv(mt.get("income_investments", "")), _sv(mt.get("income_rental", "")), _sv(mt.get("income_rollup", "")),
            _sv(mt.get("av_used", "")), _sv(mt.get("mp_used", "")),
        ])

    # ── Sheet 3: Transaction Details ──────────────────────────────
    ws3 = wb.create_sheet("3. Transaction Details")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    _set_header_row(ws3, [
        "Scheme Name", "Agency",
        "HOMES Portal (CorpPass)", "No. Orgs CorpPass",
        "HOMES Portal (Intranet)", "No. Orgs Intranet",
        "Batch SFTP", "Realtime APEX",
        "Scheduled Reports Delivery", "Required Scheduled Reports",
        "Systems via Intranet SFTP", "Systems via Intranet API",
        "Systems via Internet SFTP", "Systems via Internet API",
        "Interfacing System Status", "Interface Ready Date", "Interface System Names",
        "Annual MT Applications",
        *[f"Month {m}" for m in months],
        "Manual Recon %", "Max Concurrent Users/hr", "Reconciliation Breakdown",
        "SFTP Peak Vol", "SFTP Avg Vol",
        "API Peak Vol", "API Avg Vol",
        "Portal Peak Vol", "Portal Avg Vol",
        "Bulk Query Peak Vol", "Bulk Query Avg Vol",
    ])
    for s in submissions:
        ov = s.overview
        tx = (s.transactions.data or {}) if s.transactions else {}
        ws3.append([
            _sv(ov.scheme_name if ov else ""), _sv(ov.agency if ov else ""),
            _sv(tx.get("portal_corppass", "")), _sv(tx.get("num_orgs_corppass", "")),
            _sv(tx.get("portal_intranet", "")), _sv(tx.get("num_orgs_intranet", "")),
            _sv(tx.get("batch_sftp", "")), _sv(tx.get("realtime_apex", "")),
            _sv(tx.get("sched_reports_delivery", "")), _sv(tx.get("required_sched_reports", "")),
            _sv(tx.get("sys_intranet_sftp", "")), _sv(tx.get("sys_intranet_api", "")),
            _sv(tx.get("sys_internet_sftp", "")), _sv(tx.get("sys_internet_api", "")),
            _sv(tx.get("iface_status", "")), _sv(tx.get("iface_ready_date", "")), _sv(tx.get("iface_names", "")),
            _sv(tx.get("annual_mt_apps", "")),
            *[_sv(tx.get(f"month_{i+1}", "")) for i in range(12)],
            _sv(tx.get("manual_recon_pct", "")), _sv(tx.get("max_concurrent_users", "")), _sv(tx.get("recon_breakdown", "")),
            _sv(tx.get("vol_sftp_peak", "")), _sv(tx.get("vol_sftp_avg", "")),
            _sv(tx.get("vol_api_peak", "")), _sv(tx.get("vol_api_avg", "")),
            _sv(tx.get("vol_portal_peak", "")), _sv(tx.get("vol_portal_avg", "")),
            _sv(tx.get("vol_bulk_query_peak", "")), _sv(tx.get("vol_bulk_query_avg", "")),
        ])

    # ── Sheet 4: HOMES Functions ──────────────────────────────────
    ws4 = wb.create_sheet("4. HOMES Functions")
    events = [
        "EV001", "EV002", "EV003", "EV004", "EV005", "EV006", "EV007",
        "EV008", "EV009", "EV010", "EV011", "EV012", "EV013", "EV014",
    ]
    _set_header_row(ws4, [
        "Scheme Name", "Agency",
        "MTC View: Related", "MTC View: Nuclear", "MTC View: Parent-Guardian", "MTC View: IFM",
        "Allow Others to View?", "Can View Others?",
        "Result View: Related", "Result View: Nuclear", "Result View: Parent-Guardian", "Result View: IFM",
        "Affiliated Schemes", "Read MSHL?",
        "Subscribe Notifications?", "Beneficiary Scope",
        *[f"Event {e}" for e in events],
        "Auto-MT Subscription", "Cohort Basis",
    ])
    mtcList = ["related", "nuclear", "parent_guardian", "ifm"]
    for s in submissions:
        ov = s.overview
        hf = (s.homes_functions.data or {}) if s.homes_functions else {}
        ws4.append([
            _sv(ov.scheme_name if ov else ""), _sv(ov.agency if ov else ""),
            *[_sv(hf.get(f"view_{m}", "")) for m in mtcList],
            _sv(hf.get("allow_others_view", "")), _sv(hf.get("can_view_others", "")),
            *[_sv(hf.get(f"result_{m}", "")) for m in mtcList],
            _sv(hf.get("affiliated_schemes", "")), _sv(hf.get("read_mshl", "")),
            _sv(hf.get("subscribe_notifications", "")), _sv(hf.get("beneficiary_scope", "")),
            *[_sv(hf.get(f"event_{e}", "")) for e in events],
            _sv(hf.get("auto_mt_sub", "")), _sv(hf.get("cohort_basis", "")),
        ])

    # ── Sheet 5: MT Bands ─────────────────────────────────────────
    ws5 = wb.create_sheet("5. MT Bands")
    _set_header_row(ws5, [
        "Scheme Name", "Agency", "Row Type",
        "Band Name", "Band Version", "GI Subsidy Formula", "PCI Subsidy Formula",
        "Min Income", "Max Income", "Min AV", "Max AV",
        "Ranking Label", "Ranking Order",
    ])
    for s in submissions:
        ov = s.overview
        mb = (s.mt_bands.data or {}) if s.mt_bands else {}
        bands = mb.get("bands", []) if isinstance(mb, dict) else []
        rankings = mb.get("rankings", []) if isinstance(mb, dict) else []
        if not bands and not rankings:
            ws5.append([_sv(ov.scheme_name if ov else ""), _sv(ov.agency if ov else ""), "", "", "", "", "", "", "", "", "", "", ""])
        for b in bands:
            ws5.append([
                _sv(ov.scheme_name if ov else ""), _sv(ov.agency if ov else ""), "Band",
                _sv(b.get("band_name", "")), _sv(b.get("band_version", "")),
                _sv(b.get("gi_formula", "")), _sv(b.get("pci_formula", "")),
                _sv(b.get("min_income", "")), _sv(b.get("max_income", "")),
                _sv(b.get("min_av", "")), _sv(b.get("max_av", "")),
                "", "",
            ])
        for r in rankings:
            ws5.append([
                _sv(ov.scheme_name if ov else ""), _sv(ov.agency if ov else ""), "Ranking",
                "", "", "", "", "", "", "", "",
                _sv(r.get("label", "")), _sv(r.get("order", "")),
            ])

    # ── Sheet 6: API & Batch Interfaces ───────────────────────────
    ws6 = wb.create_sheet("6. API & Batch Interfaces")
    api_keys = [
        *[f"P12 API{i}" for i in range(1, 20)],
        "P20 API",
    ]
    hdr6 = ["Scheme Name", "Agency"]
    for k in api_keys:
        short = k.replace(" ", "").replace("P12", "").replace("P20", "P20_")
        hdr6 += [f"{k} Used", f"{k} Avg TPS", f"{k} Peak TPS"]
    _set_header_row(ws6, hdr6)
    for s in submissions:
        ov = s.overview
        ai = (s.api_interfaces.data or {}) if s.api_interfaces else {}
        row6 = [_sv(ov.scheme_name if ov else ""), _sv(ov.agency if ov else "")]
        for i in range(1, 20):
            row6 += [_sv(ai.get(f"api_P12_API{i}_used", "")), _sv(ai.get(f"api_P12_API{i}_avg", "")), _sv(ai.get(f"api_P12_API{i}_peak", ""))]
        row6 += [_sv(ai.get("api_P20_API_used", "")), _sv(ai.get("api_P20_API_avg", "")), _sv(ai.get("api_P20_API_peak", ""))]
        ws6.append(row6)

    # Auto-size columns (approximate)
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/export-bulk")
async def export_bulk(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """Export all visible schemes to a structured 6-tab Excel workbook."""
    query = (
        select(SchemeSubmission)
        .options(
            selectinload(SchemeSubmission.overview),
            selectinload(SchemeSubmission.mt_parameters),
            selectinload(SchemeSubmission.transactions),
            selectinload(SchemeSubmission.homes_functions),
            selectinload(SchemeSubmission.mt_bands),
            selectinload(SchemeSubmission.api_interfaces),
        )
        .order_by(SchemeSubmission.created_at.desc())
    )
    if not user.is_admin() and user.agency:
        query = query.join(SchemeOverview, SchemeSubmission.scheme_overview_id == SchemeOverview.id).where(SchemeOverview.agency == user.agency)
    result = await db.execute(query)
    submissions = result.scalars().all()
    buf = _build_scheme_excel(submissions)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="schemes_export.xlsx"'},
    )


# ── Required fields per tab (for import validation report) ──────
_REQUIRED_OVERVIEW = ["scheme_name", "agency", "scheme_code", "legislated_or_consent"]
_REQUIRED_MT = ["same_as_applicant", "residence_status"]
_REQUIRED_TX = ["portal_corppass", "batch_sftp", "annual_mt_apps"]
_REQUIRED_HF = ["subscribe_notifications", "auto_mt_sub"]
_REQUIRED_BANDS = ["bands"]  # at least one band must exist
_REQUIRED_API = []  # no hard required fields


def _validate_imported_scheme(name: str, agency: str, tab_data: dict) -> list[str]:
    """Return a list of human-readable missing/incomplete field warnings."""
    warnings = []
    ov = tab_data.get("overview", {})
    for f in _REQUIRED_OVERVIEW:
        if not ov.get(f):
            warnings.append(f"Overview: '{f}' is required")
    mt = tab_data.get("mt_parameters", {})
    for f in _REQUIRED_MT:
        if not mt.get(f):
            warnings.append(f"MT Parameters: '{f}' is required")
    tx = tab_data.get("transactions", {})
    for f in _REQUIRED_TX:
        if not tx.get(f):
            warnings.append(f"Transaction Details: '{f}' is required")
    hf = tab_data.get("homes_functions", {})
    for f in _REQUIRED_HF:
        if not hf.get(f):
            warnings.append(f"HOMES Functions: '{f}' is required")
    mb = tab_data.get("mt_bands", {})
    if not mb.get("bands"):
        warnings.append("MT Bands: at least one band row is required")
    return warnings


@router.post("/import-schemes", status_code=200)
async def import_schemes(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Import schemes from a structured Excel file (same 6-tab format as export).
    Bypasses validation but returns a per-scheme report of missing/incomplete fields.
    Agency-scoped: non-admin users can only import into their own agency.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are supported")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    # ── Parse each sheet into dicts keyed by (scheme_name, agency) ──
    def _ws_rows(*sheet_names: str):
        ws = None
        for n in sheet_names:
            if n in wb.sheetnames:
                ws = wb[n]
                break
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]

    def _v(row, *keys):
        for k in keys:
            if k in row and row[k] not in (None, ""):
                return str(row[k]).strip()
        return ""

    # ── Sheet 1: Overview ──
    overview_rows = _ws_rows("1. Scheme Overview", "Scheme Overview")
    schemes_map: dict[tuple, dict] = {}  # (name, agency) -> tab_data
    for row in overview_rows:
        name = _v(row, "Scheme Name")
        agency = _v(row, "Agency")
        if not name:
            continue
        key = (name, agency)
        schemes_map[key] = {
            "overview": {
                "scheme_name": name,
                "agency": agency,
                "scheme_code": _v(row, "Scheme Code"),
                "legislated_or_consent": _v(row, "Legislated/Consent"),
                "consent_scope": _v(row, "Consent Scope"),
                "valid_from": _v(row, "Valid From"),
                "valid_to": _v(row, "Valid To"),
                "background_info": {
                    "org_established": _v(row, "Org Established"),
                    "purpose": _v(row, "Purpose"),
                    "funding_source": _v(row, "Funding Source"),
                    "governing_body": _v(row, "Governing Body"),
                    "eligibility_org": _v(row, "Eligibility Org"),
                    "eval_orgs": _v(row, "Evaluating Orgs"),
                    "third_parties": _v(row, "Third Parties"),
                    "group_name": _v(row, "Group Name"),
                    "logo_info": _v(row, "Logo Info"),
                },
            },
            "mt_parameters": {},
            "transactions": {},
            "homes_functions": {},
            "mt_bands": {"bands": [], "rankings": []},
            "api_interfaces": {},
        }

    # ── Sheet 2: MT Parameters ──
    for row in _ws_rows("2. Scheme MT Parameters", "Scheme MT Parameters"):
        key = (_v(row, "Scheme Name"), _v(row, "Agency"))
        if key not in schemes_map:
            continue
        def _inclusion(col):
            raw = _v(row, col)
            if not raw:
                return {}
            try:
                parsed = _json.loads(raw)
                return parsed if isinstance(parsed, dict) else {}
            except Exception:
                return {}
        schemes_map[key]["mt_parameters"] = {
            "same_as_applicant": _v(row, "Same As Applicant"),
            "relationship_desc": _v(row, "Relationship Desc"),
            "residence_status": _v(row, "Residence Status"),
            "foreigner_pass_types": _v(row, "Foreigner Pass Types"),
            "related_used": _v(row, "Related Used"), "related_construct": _v(row, "Related Construct"), "related_deviation": _v(row, "Related Deviation"),
            "related": {"inclusion": _inclusion("Related Inclusion JSON")},
            "nuclear_used": _v(row, "Nuclear Used"), "nuclear_construct": _v(row, "Nuclear Construct"), "nuclear_deviation": _v(row, "Nuclear Deviation"),
            "nuclear": {"inclusion": _inclusion("Nuclear Inclusion JSON")},
            "parent_guardian_used": _v(row, "Parent Guardian Used"), "parent_guardian_construct": _v(row, "Parent Guardian Construct"), "parent_guardian_deviation": _v(row, "Parent Guardian Deviation"),
            "parent_guardian": {"inclusion": _inclusion("Parent Guardian Inclusion JSON")},
            "immediate_family_used": _v(row, "Immediate Family Used"), "immediate_family_construct": _v(row, "Immediate Family Construct"), "immediate_family_deviation": _v(row, "Immediate Family Deviation"),
            "immediate_family": {"inclusion": _inclusion("Immediate Family Inclusion JSON")},
            "freeform_used": _v(row, "Freeform Used"), "freeform_construct": _v(row, "Freeform Construct"), "freeform_deviation": _v(row, "Freeform Deviation"),
            "freeform": {"inclusion": _inclusion("Freeform Inclusion JSON")},
            "income_employment": _v(row, "Income Employment"), "income_trade": _v(row, "Income Trade"),
            "income_investments": _v(row, "Income Investments"), "income_rental": _v(row, "Income Rental"), "income_rollup": _v(row, "Income Rollup"),
            "av_used": _v(row, "AV Used"), "mp_used": _v(row, "MP Used"),
        }

    # ── Sheet 3: Transaction Details ──
    months_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for row in _ws_rows("3. Transaction Details", "Transaction Details"):
        key = (_v(row, "Scheme Name"), _v(row, "Agency"))
        if key not in schemes_map:
            continue
        tx: dict = {
            "portal_corppass": _v(row, "HOMES Portal (CorpPass)"), "num_orgs_corppass": _v(row, "No. Orgs CorpPass"),
            "portal_intranet": _v(row, "HOMES Portal (Intranet)"), "num_orgs_intranet": _v(row, "No. Orgs Intranet"),
            "batch_sftp": _v(row, "Batch SFTP"), "realtime_apex": _v(row, "Realtime APEX"),
            "sched_reports_delivery": _v(row, "Scheduled Reports Delivery"), "required_sched_reports": _v(row, "Required Scheduled Reports"),
            "sys_intranet_sftp": _v(row, "Systems via Intranet SFTP"), "sys_intranet_api": _v(row, "Systems via Intranet API"),
            "sys_internet_sftp": _v(row, "Systems via Internet SFTP"), "sys_internet_api": _v(row, "Systems via Internet API"),
            "iface_status": _v(row, "Interfacing System Status"), "iface_ready_date": _v(row, "Interface Ready Date"), "iface_names": _v(row, "Interface System Names"),
            "annual_mt_apps": _v(row, "Annual MT Applications"),
            "manual_recon_pct": _v(row, "Manual Recon %"), "max_concurrent_users": _v(row, "Max Concurrent Users/hr"), "recon_breakdown": _v(row, "Reconciliation Breakdown"),
            "vol_sftp_peak": _v(row, "SFTP Peak Vol"), "vol_sftp_avg": _v(row, "SFTP Avg Vol"),
            "vol_api_peak": _v(row, "API Peak Vol"), "vol_api_avg": _v(row, "API Avg Vol"),
            "vol_portal_peak": _v(row, "Portal Peak Vol"), "vol_portal_avg": _v(row, "Portal Avg Vol"),
            "vol_bulk_query_peak": _v(row, "Bulk Query Peak Vol"), "vol_bulk_query_avg": _v(row, "Bulk Query Avg Vol"),
        }
        for i, m in enumerate(months_abbr):
            tx[f"month_{i+1}"] = _v(row, f"Month {m}")
        schemes_map[key]["transactions"] = tx

    # ── Sheet 4: HOMES Functions ──
    hf_events = [f"EV{str(i).zfill(3)}" for i in range(1, 15)]
    hf_mtcs = ["related", "nuclear", "parent_guardian", "ifm"]
    for row in _ws_rows("4. HOMES Functions", "HOMES Functions"):
        key = (_v(row, "Scheme Name"), _v(row, "Agency"))
        if key not in schemes_map:
            continue
        hf: dict = {
            "allow_others_view": _v(row, "Allow Others to View?"), "can_view_others": _v(row, "Can View Others?"),
            "affiliated_schemes": _v(row, "Affiliated Schemes"), "read_mshl": _v(row, "Read MSHL?"),
            "subscribe_notifications": _v(row, "Subscribe Notifications?"), "beneficiary_scope": _v(row, "Beneficiary Scope"),
            "auto_mt_sub": _v(row, "Auto-MT Subscription"), "cohort_basis": _v(row, "Cohort Basis"),
        }
        for m in hf_mtcs:
            label = "IFM" if m == "ifm" else m.replace("_", "-").title()
            hf[f"view_{m}"] = _v(row, f"MTC View: {label}")
            hf[f"result_{m}"] = _v(row, f"Result View: {label}")
        for e in hf_events:
            hf[f"event_{e}"] = _v(row, f"Event {e}")
        schemes_map[key]["homes_functions"] = hf

    # ── Sheet 5: MT Bands (multiple rows per scheme) ──
    for row in _ws_rows("5. MT Bands", "MT Bands"):
        key = (_v(row, "Scheme Name"), _v(row, "Agency"))
        if key not in schemes_map:
            continue
        row_type = _v(row, "Row Type")
        if row_type == "Band":
            schemes_map[key]["mt_bands"]["bands"].append({
                "band_name": _v(row, "Band Name"), "band_version": _v(row, "Band Version"),
                "gi_formula": _v(row, "GI Subsidy Formula"), "pci_formula": _v(row, "PCI Subsidy Formula"),
                "min_income": _v(row, "Min Income"), "max_income": _v(row, "Max Income"),
                "min_av": _v(row, "Min AV"), "max_av": _v(row, "Max AV"),
            })
        elif row_type == "Ranking":
            schemes_map[key]["mt_bands"]["rankings"].append({
                "label": _v(row, "Ranking Label"), "order": _v(row, "Ranking Order"),
            })

    # ── Sheet 6: API & Batch Interfaces ──
    for row in _ws_rows("6. API & Batch Interfaces", "API & Batch Interfaces"):
        key = (_v(row, "Scheme Name"), _v(row, "Agency"))
        if key not in schemes_map:
            continue
        ai: dict = {}
        for i in range(1, 20):
            api_key = f"P12 API{i}"
            ai[f"api_P12_API{i}_used"] = _v(row, f"{api_key} Used")
            ai[f"api_P12_API{i}_avg"] = _v(row, f"{api_key} Avg TPS")
            ai[f"api_P12_API{i}_peak"] = _v(row, f"{api_key} Peak TPS")
        ai["api_P20_API_used"] = _v(row, "P20 API Used")
        ai["api_P20_API_avg"] = _v(row, "P20 API Avg TPS")
        ai["api_P20_API_peak"] = _v(row, "P20 API Peak TPS")
        schemes_map[key]["api_interfaces"] = ai

    # ── Persist: upsert each scheme as a draft ──────────────────
    agency_filter = None if user.is_admin() else user.agency
    results = []
    for (name, agency), tab_data in schemes_map.items():
        # Agency scope enforcement
        if agency_filter and agency != agency_filter:
            results.append({"scheme_name": name, "agency": agency, "status": "skipped", "reason": "Agency mismatch — you can only import into your own agency", "warnings": []})
            continue

        ov_data = tab_data["overview"]

        # Find existing draft or create new
        existing_q = await db.execute(
            select(SchemeSubmission)
            .join(SchemeOverview, SchemeSubmission.scheme_overview_id == SchemeOverview.id)
            .where(SchemeOverview.scheme_name == name, SchemeOverview.agency == agency, SchemeSubmission.status == SubmissionStatus.draft.value)
            .options(
                selectinload(SchemeSubmission.overview),
                selectinload(SchemeSubmission.mt_parameters),
                selectinload(SchemeSubmission.transactions),
                selectinload(SchemeSubmission.homes_functions),
                selectinload(SchemeSubmission.mt_bands),
                selectinload(SchemeSubmission.api_interfaces),
            )
            .limit(1)
        )
        sub = existing_q.scalars().first()
        imported_new = sub is None

        if sub is None:
            # Create new master + overview + submission
            master = SchemeMaster(agency=agency or "", scheme_name=name, scheme_code=ov_data.get("scheme_code"), created_by=user.id)
            db.add(master)
            await db.flush()

            ov = SchemeOverview(
                agency=agency, scheme_name=name, scheme_code=ov_data.get("scheme_code"),
                legislated_or_consent=ov_data.get("legislated_or_consent"),
                consent_scope=ov_data.get("consent_scope"),
                background_info=ov_data.get("background_info", {}),
            )
            db.add(ov)
            await db.flush()

            valid_from = None
            valid_to = None
            try:
                if ov_data.get("valid_from"):
                    valid_from = date.fromisoformat(str(ov_data["valid_from"])[:10])
                if ov_data.get("valid_to"):
                    valid_to = date.fromisoformat(str(ov_data["valid_to"])[:10])
            except Exception:
                pass

            sub = SchemeSubmission(
                scheme_master_id=master.id, scheme_overview_id=ov.id,
                status=SubmissionStatus.draft.value, version=1,
                valid_from=valid_from, valid_to=valid_to, created_by=user.id,
            )
            db.add(sub)
            await db.flush()
        else:
            # Update existing overview
            ov = sub.overview
            if ov:
                ov.scheme_code = ov_data.get("scheme_code") or ov.scheme_code
                ov.legislated_or_consent = ov_data.get("legislated_or_consent") or ov.legislated_or_consent
                ov.consent_scope = ov_data.get("consent_scope") or ov.consent_scope
                ov.background_info = ov_data.get("background_info") or ov.background_info

        # Upsert each tab's JSON data
        async def _upsert_tab(model_cls, rel_name: str, data_key: str):
            obj_res = await db.execute(
                select(model_cls).where(model_cls.submission_id == sub.id).limit(1)
            )
            obj = obj_res.scalars().first()
            data = tab_data.get(data_key, {})
            if obj is None:
                obj = model_cls(submission_id=sub.id, data=data)
                db.add(obj)
            else:
                obj.data = data

        await _upsert_tab(SchemeMTParameters, "mt_parameters", "mt_parameters")
        await _upsert_tab(TransactionDetails, "transactions", "transactions")
        await _upsert_tab(HOMESFunctions, "homes_functions", "homes_functions")
        await _upsert_tab(MTBands, "mt_bands", "mt_bands")
        await _upsert_tab(APIBatchInterfaces, "api_interfaces", "api_interfaces")

        # Validate and collect warnings
        warnings = _validate_imported_scheme(name, agency, tab_data)

        results.append({
            "scheme_name": name,
            "agency": agency,
            "status": "created" if imported_new else "updated",
            "warnings": warnings,
        })

    await db.commit()
    return {"imported": len([r for r in results if r["status"] in ("created", "updated")]), "results": results}

@router.get("")
async def list_schemes(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    query = (
        select(SchemeSubmission)
        .options(selectinload(SchemeSubmission.master), selectinload(SchemeSubmission.overview), selectinload(SchemeSubmission.creator), selectinload(SchemeSubmission.onboarding_slots))
        .order_by(SchemeSubmission.created_at.desc())
    )
    # Agency-scoped: non-admin users only see their own agency's schemes
    if not user.is_admin() and user.agency:
        query = query.join(SchemeOverview, SchemeSubmission.scheme_overview_id == SchemeOverview.id).where(SchemeOverview.agency == user.agency)
    result = await db.execute(query)
    subs = result.scalars().all()
    return [
        {
            "id": s.id,
            "scheme_master_id": s.scheme_master_id,
            "scheme_name": s.master.scheme_name if s.master else (s.overview.scheme_name if s.overview else None),
            "scheme_code": s.master.scheme_code if s.master else (s.overview.scheme_code if s.overview else None),
            "agency": s.master.agency if s.master else (s.overview.agency if s.overview else None),
            "status": _effective_status(s),
            "version": s.version,
            "version_label": f"v{s.version}",
            "valid_from": s.valid_from.isoformat() if s.valid_from else None,
            "valid_to": s.valid_to.isoformat() if s.valid_to else None,
            "created_by": s.creator.display_name if s.creator else None,
            "updated_at": s.updated_at.isoformat() if s.updated_at else None,
            "primary_slot": next(
                ({"year": slot.year, "slot_month": slot.slot_month, "slot_month_name": slot.slot_month_name}
                 for slot in (s.onboarding_slots or []) if not slot.is_additional),
                None
            ),
        }
        for s in subs
    ]


@router.post("", status_code=201)
async def create_scheme(body: SchemeCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "agency_creator")
    if body.agency and body.agency != user.agency:
        raise HTTPException(status_code=403, detail="Creators can only create schemes for their own agency")
    valid_from = _parse_iso_date(body.valid_from, "valid_from")
    valid_to = _parse_iso_date(body.valid_to, "valid_to")
    if valid_from and valid_to and valid_from > valid_to:
        raise HTTPException(status_code=400, detail="valid_from cannot be after valid_to")

    # Find or create master record (Agency + Scheme Name + Scheme Code)
    agency_val = body.agency or user.agency
    master_res = await db.execute(
        select(SchemeMaster).where(
            SchemeMaster.agency == agency_val,
            SchemeMaster.scheme_name == body.scheme_name,
            SchemeMaster.scheme_code == body.scheme_code,
        )
    )
    master = master_res.scalar_one_or_none()
    if not master:
        master = SchemeMaster(
            agency=agency_val or "",
            scheme_name=body.scheme_name,
            scheme_code=body.scheme_code,
            created_by=user.id,
        )
        db.add(master)
        await db.flush()

    vres = await db.execute(select(SchemeSubmission).where(SchemeSubmission.scheme_master_id == master.id))
    existing_versions = vres.scalars().all()
    next_version = (max([v.version for v in existing_versions], default=0) + 1)
    await _validate_version_window(db, master.id, valid_from, valid_to)

    overview = SchemeOverview(
        agency=agency_val,
        scheme_name=master.scheme_name,
        scheme_code=master.scheme_code,
        legislated_or_consent=body.legislated_or_consent,
        consent_scope=body.consent_scope,
        background_info=body.background_info,
    )
    db.add(overview)
    await db.flush()

    submission = SchemeSubmission(
        scheme_master_id=master.id,
        scheme_overview_id=overview.id,
        status=SubmissionStatus.draft.value,
        version=next_version,
        valid_from=valid_from,
        valid_to=valid_to,
        created_by=user.id,
    )
    db.add(submission)
    await db.commit()
    await db.refresh(submission)
    return {
        "id": submission.id,
        "scheme_master_id": master.id,
        "scheme_name": overview.scheme_name,
        "status": submission.status,
        "version": submission.version,
        "version_label": f"v{submission.version}",
    }


@router.get("/{scheme_id}")
async def get_scheme(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    return _sub_to_dict(sub)


@router.put("/{scheme_id}")
async def update_scheme(scheme_id: str, body: SchemeUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    await _ensure_master_for_submission(db, sub, user)
    _assert_agency_access(user, sub)
    if not _can_edit_submission(user, sub):
        raise HTTPException(400, "You are not allowed to edit this scheme in its current status")

    ov = sub.overview
    requested_updates: dict = {}
    for field in ("agency", "scheme_name", "scheme_code", "legislated_or_consent", "consent_scope", "background_info"):
        new_val = getattr(body, field, None)
        if new_val is not None:
            requested_updates[field] = new_val

    new_valid_from = _parse_iso_date(body.valid_from, "valid_from") if body.valid_from is not None else sub.valid_from
    new_valid_to = _parse_iso_date(body.valid_to, "valid_to") if body.valid_to is not None else sub.valid_to
    await _validate_version_window(db, sub.scheme_master_id, new_valid_from, new_valid_to, exclude_submission_id=sub.id)

    old_snapshot = {field: getattr(ov, field) for field in requested_updates.keys()}
    old_snapshot["valid_from"] = sub.valid_from.isoformat() if sub.valid_from else None
    old_snapshot["valid_to"] = sub.valid_to.isoformat() if sub.valid_to else None
    new_snapshot = dict(requested_updates)
    new_snapshot["valid_from"] = new_valid_from.isoformat() if new_valid_from else None
    new_snapshot["valid_to"] = new_valid_to.isoformat() if new_valid_to else None
    changes = _deep_field_diff(old_snapshot, new_snapshot)

    for field, value in requested_updates.items():
        setattr(ov, field, value)
    sub.valid_from = new_valid_from
    sub.valid_to = new_valid_to

    if sub.master:
        if "agency" in requested_updates and requested_updates["agency"] is not None:
            sub.master.agency = requested_updates["agency"]
        if "scheme_name" in requested_updates and requested_updates["scheme_name"] is not None:
            sub.master.scheme_name = requested_updates["scheme_name"]
        if "scheme_code" in requested_updates:
            sub.master.scheme_code = requested_updates["scheme_code"]

    if changes:
        db.add(ChangeLog(submission_id=sub.id, changed_by=user.id, tab_name="overview", changes=changes))

    sub.updated_at = sg_now()
    await db.commit()
    return {"ok": True, "changes": len(changes)}


@router.put("/{scheme_id}/tab/{tab_name}")
async def update_tab(scheme_id: str, tab_name: str, body: TabUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    if tab_name not in TAB_MODEL_MAP:
        raise HTTPException(400, f"Invalid tab: {tab_name}. Must be one of {list(TAB_MODEL_MAP.keys())}")

    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    if not _can_edit_submission(user, sub):
        raise HTTPException(400, "You are not allowed to edit this scheme in its current status")

    model_cls = TAB_MODEL_MAP[tab_name]
    rel_name = TAB_REL_MAP[tab_name]
    existing = getattr(sub, rel_name)

    old_data = (existing.data if existing and existing.data is not None else {})
    new_data = body.data or {}
    changes = _deep_field_diff(old_data, new_data)

    if existing:
        existing.data = new_data
    else:
        new_obj = model_cls(submission_id=sub.id, data=new_data)
        db.add(new_obj)

    if changes:
        db.add(ChangeLog(submission_id=sub.id, changed_by=user.id, tab_name=tab_name, changes=changes))
    sub.updated_at = sg_now()
    await db.commit()
    return {"ok": True, "tab": tab_name, "changes": len(changes)}


@router.get("/{scheme_id}/versions")
async def list_versions(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    await _ensure_master_for_submission(db, sub, user)
    _assert_agency_access(user, sub)

    res = await db.execute(
        select(SchemeSubmission)
        .options(selectinload(SchemeSubmission.master), selectinload(SchemeSubmission.overview))
        .where(SchemeSubmission.scheme_master_id == sub.scheme_master_id)
        .order_by(SchemeSubmission.version.asc())
    )
    versions = res.scalars().all()
    return [
        {
            "id": v.id,
            "version": v.version,
            "version_label": f"v{v.version}",
            "status": _effective_status(v),
            "valid_from": v.valid_from.isoformat() if v.valid_from else None,
            "valid_to": v.valid_to.isoformat() if v.valid_to else None,
        }
        for v in versions
    ]


@router.post("/{scheme_id}/clone-version", status_code=201)
async def clone_version(scheme_id: str, body: CloneVersionBody, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "agency_creator", "mto_admin")
    source = await _get_submission(db, scheme_id)
    await _ensure_master_for_submission(db, source, user)
    _assert_agency_access(user, source)

    vf = _parse_iso_date(body.valid_from, "valid_from")
    vt = _parse_iso_date(body.valid_to, "valid_to")
    await _validate_version_window(db, source.scheme_master_id, vf, vt)

    vers_res = await db.execute(select(SchemeSubmission).where(SchemeSubmission.scheme_master_id == source.scheme_master_id))
    existing = vers_res.scalars().all()
    next_version = max([x.version for x in existing], default=0) + 1

    src_ov = source.overview
    overview = SchemeOverview(
        agency=src_ov.agency if src_ov else (source.master.agency if source.master else user.agency),
        scheme_name=src_ov.scheme_name if src_ov else (source.master.scheme_name if source.master else ""),
        scheme_code=src_ov.scheme_code if src_ov else (source.master.scheme_code if source.master else None),
        legislated_or_consent=src_ov.legislated_or_consent if src_ov else None,
        consent_scope=src_ov.consent_scope if src_ov else None,
        background_info=(dict(src_ov.background_info) if src_ov and src_ov.background_info else None),
    )
    db.add(overview)
    await db.flush()

    clone = SchemeSubmission(
        scheme_master_id=source.scheme_master_id,
        scheme_overview_id=overview.id,
        status=SubmissionStatus.draft.value,
        version=next_version,
        valid_from=vf,
        valid_to=vt,
        cloned_from_submission_id=source.id,
        created_by=user.id,
    )
    db.add(clone)
    await db.flush()

    def clone_tab(existing_obj, model_cls):
        if not existing_obj:
            return
        copied = dict(existing_obj.data) if existing_obj.data else {}
        db.add(model_cls(submission_id=clone.id, data=copied))

    clone_tab(source.mt_parameters, SchemeMTParameters)
    clone_tab(source.transactions, TransactionDetails)
    clone_tab(source.homes_functions, HOMESFunctions)
    clone_tab(source.mt_bands, MTBands)
    clone_tab(source.api_interfaces, APIBatchInterfaces)

    db.add(Comment(submission_id=clone.id, user_id=user.id, text=f"Cloned from v{source.version}", stage="cloned"))
    await db.commit()
    return {"id": clone.id, "version": clone.version, "version_label": f"v{clone.version}", "status": clone.status}


@router.post("/{scheme_id}/activate")
async def activate_version(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "mto_admin")
    sub = await _get_submission(db, scheme_id)
    await _ensure_master_for_submission(db, sub, user)
    if sub.status != SubmissionStatus.approved.value:
        raise HTTPException(status_code=400, detail=f"Only approved versions can be activated (current: {sub.status})")
    await _validate_version_window(db, sub.scheme_master_id, sub.valid_from, sub.valid_to, exclude_submission_id=sub.id)

    res = await db.execute(select(SchemeSubmission).where(SchemeSubmission.scheme_master_id == sub.scheme_master_id))
    versions = res.scalars().all()
    for v in versions:
        if v.id == sub.id:
            continue
        if _effective_status(v) == SubmissionStatus.active.value:
            v.status = SubmissionStatus.expired.value

    sub.status = SubmissionStatus.active.value
    sub.updated_at = sg_now()
    db.add(Comment(submission_id=sub.id, user_id=user.id, text="Version activated", stage="active"))
    await db.commit()
    return {"ok": True, "status": sub.status}


@router.post("/{scheme_id}/retire")
async def retire_version(scheme_id: str, body: RetireBody = RetireBody(), db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "mto_admin")
    sub = await _get_submission(db, scheme_id)
    if _effective_status(sub) != SubmissionStatus.active.value:
        raise HTTPException(status_code=400, detail="Only active versions can be retired")

    retire_date = _parse_iso_date(body.retire_date, "retire_date") if body.retire_date else date.today()
    sub.valid_to = retire_date
    sub.status = SubmissionStatus.retired.value
    sub.updated_at = sg_now()
    db.add(Comment(submission_id=sub.id, user_id=user.id, text=(body.comment or "Version retired by MTO"), stage="retired"))
    await db.commit()
    return {"ok": True, "status": sub.status, "valid_to": sub.valid_to.isoformat() if sub.valid_to else None}


# ── Workflow ─────────────────────────────────────────────────────

@router.post("/{scheme_id}/submit")
async def submit_scheme(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "agency_creator")
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    if sub.status not in (SubmissionStatus.draft.value, SubmissionStatus.rejected.value):
        raise HTTPException(400, f"Cannot submit from status {sub.status}")
    sub.status = SubmissionStatus.pending_review.value
    sub.updated_at = sg_now()
    db.add(Comment(submission_id=sub.id, user_id=user.id, text="Submitted for agency review", stage="submitted"))
    await db.commit()
    await _notify_for_workflow_stage(db, sub, SubmissionStatus.pending_review.value, triggered_by=user.id)
    return {"ok": True, "status": sub.status}


@router.post("/{scheme_id}/approve")
async def approve_scheme(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "agency_approver")
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    if sub.status not in (SubmissionStatus.pending_review.value, SubmissionStatus.rejected.value):
        raise HTTPException(400, f"Cannot approve/send from status {sub.status}")
    sub.status = SubmissionStatus.pending_final.value
    sub.updated_at = sg_now()
    db.add(Comment(submission_id=sub.id, user_id=user.id, text="Approved by agency approver and sent to MTO", stage="approved"))
    await db.commit()
    await _notify_for_workflow_stage(db, sub, SubmissionStatus.pending_final.value, triggered_by=user.id)
    return {"ok": True, "status": sub.status}


@router.post("/{scheme_id}/final-approve")
async def final_approve_scheme(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "mto_admin")
    sub = await _get_submission(db, scheme_id)
    if sub.status != SubmissionStatus.pending_final.value:
        raise HTTPException(400, f"Cannot final-approve from status {sub.status}")
    sub.status = SubmissionStatus.approved.value
    sub.updated_at = sg_now()
    db.add(Comment(submission_id=sub.id, user_id=user.id, text="Final approval granted", stage="approved"))
    await db.commit()
    return {"ok": True, "status": sub.status}


@router.post("/{scheme_id}/reject")
async def reject_scheme(scheme_id: str, body: RejectBody = RejectBody(), db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)

    if user.has_role("agency_approver"):
        if sub.status != SubmissionStatus.pending_review.value:
            raise HTTPException(400, f"Agency approver cannot reject from status {sub.status}")
    elif user.has_role("mto_admin"):
        if sub.status != SubmissionStatus.pending_final.value:
            raise HTTPException(400, f"MTO admin cannot reject from status {sub.status}")
    else:
        raise HTTPException(status_code=403, detail="Only agency approver or MTO admin can reject")

    sub.status = SubmissionStatus.rejected.value
    sub.updated_at = sg_now()
    comment_text = body.comment or "Rejected"
    db.add(Comment(submission_id=sub.id, user_id=user.id, text=comment_text, stage="rejected"))
    await db.commit()
    return {"ok": True, "status": sub.status}


# ── Comments & Changes ───────────────────────────────────────────

@router.get("/{scheme_id}/comments")
async def list_comments(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    result = await db.execute(
        select(Comment).where(Comment.submission_id == scheme_id).order_by(Comment.created_at.desc())
    )
    comments = result.scalars().all()
    out = []
    for c in comments:
        u = await db.get(User, c.user_id) if c.user_id else None
        out.append({
            "id": c.id, "text": c.text, "stage": c.stage,
            "user": u.display_name if u else None,
            "created_at": c.created_at.isoformat() if c.created_at else None,
        })
    return out


@router.post("/{scheme_id}/comments", status_code=201)
async def add_comment(scheme_id: str, body: CommentCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    c = Comment(submission_id=scheme_id, user_id=user.id, text=body.text, stage=body.stage)
    db.add(c)
    await db.commit()
    return {"id": c.id, "text": c.text}


@router.get("/{scheme_id}/changes")
async def list_changes(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    result = await db.execute(
        select(ChangeLog).where(ChangeLog.submission_id == scheme_id).order_by(ChangeLog.timestamp.desc())
    )
    logs = result.scalars().all()
    out = []
    for log in logs:
        u = await db.get(User, log.changed_by) if log.changed_by else None
        out.append({
            "id": log.id, "tab_name": log.tab_name, "changes": log.changes,
            "changed_by": u.display_name if u else None,
            "changed_by_username": u.username if u else None,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
        })
    return out


# ── Excel Export ─────────────────────────────────────────────────

@router.get("/{scheme_id}/export")
async def export_excel(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    wb = Workbook()

    # Tab 1: Overview
    ws = wb.active
    ws.title = "Scheme Overview"
    ov = sub.overview
    if ov:
        for i, (k, v) in enumerate([
            ("Agency", ov.agency), ("Scheme Name", ov.scheme_name), ("Scheme Code", ov.scheme_code),
            ("Legislated/Consent", ov.legislated_or_consent), ("Consent Scope", ov.consent_scope),
        ], 1):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=str(v) if v else "")
        if ov.background_info:
            row = 7
            for k, v in ov.background_info.items():
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=str(v) if v else "")
                row += 1

    # Tabs 2-6: JSON data tabs
    tab_data = [
        ("MT Parameters", sub.mt_parameters),
        ("Transactions", sub.transactions),
        ("HOMES Functions", sub.homes_functions),
        ("MT Bands", sub.mt_bands),
        ("API Interfaces", sub.api_interfaces),
    ]
    for title, obj in tab_data:
        ws = wb.create_sheet(title)
        if obj and obj.data:
            _write_json_sheet(ws, obj.data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"scheme_{sub.overview.scheme_name if sub.overview else scheme_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_json_sheet(ws, data: dict, start_row: int = 1):
    """Flatten a dict/list into worksheet rows."""
    row = start_row
    if isinstance(data, dict):
        for k, v in data.items():
            ws.cell(row=row, column=1, value=str(k))
            if isinstance(v, (dict, list)):
                ws.cell(row=row, column=2, value=str(v))
            else:
                ws.cell(row=row, column=2, value=str(v) if v is not None else "")
            row += 1
    elif isinstance(data, list):
        for i, item in enumerate(data):
            ws.cell(row=row, column=1, value=f"Item {i+1}")
            ws.cell(row=row, column=2, value=str(item))
            row += 1


@router.delete("/{scheme_id}", status_code=200)
async def delete_scheme(scheme_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    _require_role(user, "mto_admin")
    sub = await _get_submission(db, scheme_id)
    overview = sub.overview
    master_id = sub.scheme_master_id

    await db.delete(sub)
    if overview:
        await db.delete(overview)

    if master_id:
        cnt_res = await db.execute(select(SchemeSubmission).where(SchemeSubmission.scheme_master_id == master_id, SchemeSubmission.id != scheme_id))
        remaining = cnt_res.scalars().first()
        if not remaining:
            master = await db.get(SchemeMaster, master_id)
            if master:
                await db.delete(master)
    await db.commit()
    return {"ok": True, "deleted_scheme_id": scheme_id}


# ── ONBOARDING SLOT APPROVAL ────────────────────────────────────

class SlotApprovalBody(BaseModel):
    approval_status: str  # "approved" or "rejected"
    approver_comment: str | None = None


@router.post("/{scheme_id}/slot/approve")
async def approve_slot(
    scheme_id: str,
    body: SlotApprovalBody,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Review/approve/reject the onboarding slot for a scheme submission.

    Rules:
    - Agency approver can review and reject with feedback at pending_review stage
      (cannot grant final slot approval)
    - MTO admin can approve/reject at pending_final stage
    """
    _require_role(user, "agency_approver", "mto_admin")
    
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)

    if user.has_role("agency_approver") and sub.status != SubmissionStatus.pending_review.value:
        raise HTTPException(status_code=400, detail="Agency approver can only review slot at pending_review stage")
    if user.has_role("mto_admin") and sub.status != SubmissionStatus.pending_final.value:
        raise HTTPException(status_code=400, detail="MTO admin can only review slot at pending_final stage")
    
    # Get primary slot
    primary_slot = next((s for s in (sub.onboarding_slots or []) if not s.is_additional), None)
    if not primary_slot:
        raise HTTPException(status_code=404, detail="No onboarding slot found for this submission")
    
    if body.approval_status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Status must be 'approved' or 'rejected'")

    # Final slot approval is MTO-only
    if user.has_role("agency_approver") and body.approval_status == "approved":
        raise HTTPException(status_code=403, detail="Only MTO Admin can approve onboarding slots")
    if user.has_role("agency_approver") and body.approval_status == "rejected" and not (body.approver_comment or "").strip():
        raise HTTPException(status_code=400, detail="Please provide review comments when rejecting a slot")
    
    primary_slot.approval_status = body.approval_status
    primary_slot.approver_comment = body.approver_comment
    sub.updated_at = sg_now()
    
    await db.commit()
    return {"message": f"Slot {body.approval_status}"}

# ── ONBOARDING SLOT MANAGEMENT ──────────────────────────────────

from datetime import date

def _get_month_name(month: int) -> str:
    """Convert month number to name."""
    names = {1: "January", 4: "April", 7: "July", 11: "November"}
    return names.get(month, "Unknown")


@router.put("/{scheme_id}/slot", status_code=200)
async def set_scheme_slot(
    scheme_id: str,
    body: SlotSelection,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Set or update the primary onboarding slot for a scheme submission.
    
    Only allowed for draft or rejected submissions.
    Validates:
    - Cannot select past quarters
    - Go-live dates must not be before the slot month
    - If scheme already has a slot, updates it (unless approved)
    """
    _require_role(user, "agency_creator", "agency_approver", "mto_admin")
    
    # Get submission with all slots
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    
    # Find or create primary slot
    primary_slot = next((s for s in (sub.onboarding_slots or []) if not s.is_additional), None)

    # Check slot is in editable state
    if not _can_edit_primary_slot(user, sub, primary_slot):
        raise HTTPException(status_code=400, detail="Cannot edit slot in current workflow stage")

    # Check slot month is valid (only Jan, Apr, Jul, Nov)
    if body.slot_month not in [1, 4, 7, 11]:
        raise HTTPException(status_code=400, detail="Invalid slot month. Must be January(1), April(4), July(7), or November(11)")

    # Parse dates
    try:
        tech_date = date.fromisoformat(body.technical_go_live)
        biz_date = date.fromisoformat(body.business_go_live)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # Prevent backdating: cannot select a slot in a past quarter
    today = date.today()
    slot_quarter_start = date(body.year, body.slot_month, 1)
    if slot_quarter_start < today.replace(day=1):
        raise HTTPException(status_code=400, detail=f"Cannot select a slot in the past (slot month: {_get_month_name(body.slot_month)} {body.year})")

    # Go-live dates cannot be before the selected slot month
    slot_month_date = date(body.year, body.slot_month, 1)
    if tech_date < slot_month_date or biz_date < slot_month_date:
        raise HTTPException(status_code=400, detail=f"Go-live dates cannot be before {_get_month_name(body.slot_month)} {body.year}")

    # Capacity check: max 4 bookings per quarter month (pending/approved) across all schemes
    cap_q = select(OnboardingSlot).where(
        OnboardingSlot.year == body.year,
        OnboardingSlot.slot_month == body.slot_month,
        OnboardingSlot.is_additional == False,
        OnboardingSlot.approval_status.in_(["pending", "approved"]),
    )
    existing_for_quarter = (await db.execute(cap_q)).scalars().all()
    if primary_slot:
        existing_for_quarter = [s for s in existing_for_quarter if s.id != primary_slot.id]
    if len(existing_for_quarter) >= 4:
        raise HTTPException(
            status_code=400,
            detail=f"{_get_month_name(body.slot_month)} {body.year} is full (4/4 slots). Please select another quarter.",
        )
    
    if primary_slot:
        # Update existing primary slot
        primary_slot.year = body.year
        primary_slot.slot_month = body.slot_month
        primary_slot.slot_month_name = _get_month_name(body.slot_month)
        primary_slot.technical_go_live = tech_date
        primary_slot.business_go_live = biz_date
        primary_slot.approval_status = "pending"  # Reset to pending when changed
        primary_slot.approver_comment = None
    else:
        # Create new primary slot
        primary_slot = OnboardingSlot(
            scheme_submission_id=scheme_id,
            year=body.year,
            slot_month=body.slot_month,
            slot_month_name=_get_month_name(body.slot_month),
            is_additional=False,
            technical_go_live=tech_date,
            business_go_live=biz_date,
            booked_by_id=user.id,
            approval_status="pending",
        )
        db.add(primary_slot)
    
    sub.updated_at = sg_now()
    await db.commit()
    return {"message": "Slot updated successfully"}


@router.get("/{scheme_id}/slot")
async def get_scheme_slot(
    scheme_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Get the onboarding slot(s) for a scheme submission."""
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    
    slots = []
    if sub.onboarding_slots:
        for slot in sorted(sub.onboarding_slots, key=lambda x: (x.is_additional, x.slot_month)):
            slots.append({
                "id": slot.id,
                "year": slot.year,
                "slot_month": slot.slot_month,
                "slot_month_name": slot.slot_month_name,
                "is_additional": slot.is_additional,
                "technical_go_live": slot.technical_go_live.isoformat() if slot.technical_go_live else None,
                "business_go_live": slot.business_go_live.isoformat() if slot.business_go_live else None,
                "justification": slot.justification,
                "approval_status": slot.approval_status,
                "approver_comment": slot.approver_comment,
            })
    
    return {"slots": slots}


@router.delete("/{scheme_id}/slot/{slot_id}", status_code=200)
async def delete_additional_slot(
    scheme_id: str,
    slot_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Remove a slot from a submission.

    - agency_creator/agency_approver: can remove additional slots only
    - mto_admin: can remove any slot (including primary)
    """
    _require_role(user, "agency_creator", "agency_approver", "mto_admin")
    
    sub = await _get_submission(db, scheme_id)
    _assert_agency_access(user, sub)
    
    if not _can_edit_submission(user, sub):
        raise HTTPException(status_code=400, detail="Cannot delete slot in current workflow stage")
    
    slot = next((s for s in (sub.onboarding_slots or []) if s.id == slot_id), None)
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    
    if not slot.is_additional and not user.has_role("mto_admin"):
        raise HTTPException(status_code=400, detail="Cannot delete primary slot. Use PUT to update it.")
    
    await db.delete(slot)
    sub.updated_at = sg_now()
    await db.commit()
    
    return {"message": "Slot deleted successfully"}
